import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib import admin
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.utils import timezone
from django.http import HttpResponse
from .models import Enquiry, EnquiryImage


# ── EXCEL HELPER ───────────────────────────────────────────────────────

def export_enquiries_to_excel(queryset, filename):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Enquiries"

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="CCCCCC")
    cell_border = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side,
    )
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    headers = [
        "Enquiry No.", "Order Type", "Status", "Source Page",
        "Full Name", "Phone", "Email", "Brand Name",
        "Company Age (Yrs)", "Total Pieces Req.", "Annual Revenue",
        "Message", "WL Prototype", "Fabric",
        "Assigned To", "Is Viewed", "Viewed At",
        "Admin Notes", "Created At",
    ]

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border
    ws.row_dimensions[1].height = 28

    # Data rows
    for row_num, enq in enumerate(queryset, start=2):
        ws.append([
            enq.enquiry_number,
            enq.get_order_type_display(),
            enq.get_status_display(),
            enq.get_source_page_display() if enq.source_page else "—",
            enq.full_name,
            enq.phone,
            enq.email,
            enq.brand_name,
            enq.company_age_years or "—",
            enq.total_pieces_required or "—",
            enq.annual_revenue or "—",
            enq.message,
            enq.wl_prototype.prototype_code if enq.wl_prototype else "—",
            enq.fabric.fabric_name if enq.fabric else "—",
            enq.assigned_to_user.email if enq.assigned_to_user else "—",
            "Yes" if enq.is_viewed else "No",
            enq.viewed_at.strftime("%d %b %Y %H:%M") if enq.viewed_at else "—",
            enq.admin_notes or "—",
            enq.created_at.strftime("%d %b %Y %H:%M"),
        ])
        if row_num % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = alt_fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).border = cell_border

    # Column widths
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 4, 16)

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── INLINE ─────────────────────────────────────────────────────────────

class EnquiryImageInline(admin.TabularInline):
    model           = EnquiryImage
    extra           = 0
    readonly_fields = ['id', 'image_preview', 'file_name', 'file_size_bytes', 'mime_type', 'uploaded_at']
    fields          = ['image', 'image_preview', 'file_name', 'file_size_bytes', 'mime_type', 'uploaded_at']

    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height:80px;width:80px;object-fit:cover;border-radius:4px;" />',
                obj.image.url,
            )
        return "—"
    image_preview.short_description = "Preview"


# ── ENQUIRY ADMIN ──────────────────────────────────────────────────────

@admin.register(Enquiry)
class EnquiryAdmin(admin.ModelAdmin):
    list_display = [
        'enquiry_number', 'order_type', 'full_name', 'phone',
        'email', 'brand_name', 'status', 'unread_badge',
        'assigned_to_user', 'source_page', 'created_at',
    ]
    list_filter     = ['order_type', 'status', 'source_page', 'is_viewed', 'created_at']
    search_fields   = ['enquiry_number', 'full_name', 'phone', 'email', 'brand_name']
    readonly_fields = ['id', 'enquiry_number', 'created_at', 'updated_at', 'viewed_at']
    ordering        = ['-created_at']
    inlines         = [EnquiryImageInline]
    actions         = [
        'export_all_as_excel',
        'export_private_label_as_excel',
        'export_white_label_as_excel',
        'export_fabrics_as_excel',
        'mark_as_viewed',
    ]

    fieldsets = (
        ("Enquiry Info", {
            "fields": ("id", "enquiry_number", "order_type", "source_page"),
        }),
        ("Contact Details", {
            "fields": ("full_name", "phone", "email", "brand_name",
                       "company_age_years", "annual_revenue"),
        }),
        ("Order Details", {
            "fields": ("total_pieces_required", "message"),
        }),
        ("References", {
            "fields": ("wl_prototype", "fabric", "customer"),
            "description": "WL prototype and fabric are auto-linked from the enquiry source page.",
        }),
        ("Admin Management", {
            "fields": ("status", "assigned_to_user", "admin_notes",
                       "is_viewed", "viewed_at"),
        }),
        ("Timestamps", {
            "fields": ("created_at", "updated_at"),
        }),
    )

    # ── Display methods ────────────────────────────────────────────── #

    def unread_badge(self, obj):
        if not obj.is_viewed:
            return mark_safe(
                '<span style="background:#e74c3c;color:#fff;'
                'padding:2px 8px;border-radius:10px;font-size:11px;">NEW</span>'
            )
        return "—"
    unread_badge.short_description = "Read"

    # ── Save ───────────────────────────────────────────────────────── #

    def save_model(self, request, obj, form, change):
        if change and not obj.is_viewed:
            obj.is_viewed = True
            obj.viewed_at = timezone.now()
        super().save_model(request, obj, form, change)

    # ── Actions ────────────────────────────────────────────────────── #

    @admin.action(description="📥 Export selected enquiries to Excel")
    def export_all_as_excel(self, request, queryset):
        filename = f"enquiries_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return export_enquiries_to_excel(
            queryset.select_related("assigned_to_user", "wl_prototype", "fabric"),
            filename,
        )

    @admin.action(description="📥 Export Private Label enquiries to Excel")
    def export_private_label_as_excel(self, request, queryset):
        qs = queryset.filter(order_type="private_label").select_related(
            "assigned_to_user", "wl_prototype", "fabric"
        )
        return export_enquiries_to_excel(qs, f"enquiries_pl_{timezone.now().strftime('%Y%m%d')}.xlsx")

    @admin.action(description="📥 Export White Label enquiries to Excel")
    def export_white_label_as_excel(self, request, queryset):
        qs = queryset.filter(order_type="white_label").select_related(
            "assigned_to_user", "wl_prototype", "fabric"
        )
        return export_enquiries_to_excel(qs, f"enquiries_wl_{timezone.now().strftime('%Y%m%d')}.xlsx")

    @admin.action(description="📥 Export Fabrics enquiries to Excel")
    def export_fabrics_as_excel(self, request, queryset):
        qs = queryset.filter(order_type="fabrics").select_related(
            "assigned_to_user", "wl_prototype", "fabric"
        )
        return export_enquiries_to_excel(qs, f"enquiries_fabrics_{timezone.now().strftime('%Y%m%d')}.xlsx")

    @admin.action(description="✅ Mark selected enquiries as viewed")
    def mark_as_viewed(self, request, queryset):
        updated = queryset.filter(is_viewed=False).update(
            is_viewed=True, viewed_at=timezone.now()
        )
        self.message_user(request, f"{updated} enquiry(s) marked as viewed.")


@admin.register(EnquiryImage)
class EnquiryImageAdmin(admin.ModelAdmin):
    list_display    = ['file_name', 'enquiry', 'image_preview', 'file_size_bytes', 'mime_type', 'uploaded_at']
    search_fields   = ['file_name', 'enquiry__enquiry_number']
    readonly_fields = ['id', 'image_preview', 'uploaded_at']

    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height:80px;width:80px;object-fit:cover;border-radius:4px;" />',
                obj.image.url,
            )
        return "—"
    image_preview.short_description = "Preview"