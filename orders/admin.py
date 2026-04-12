# orders/admin.py

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib import admin
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.utils import timezone
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth import get_user_model
from .models import Order, OrderStageHistory, OrderImage, OrderNote


# ── EXCEL HELPER ───────────────────────────────────────────────────────

def export_orders_to_excel(queryset, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="CCCCCC")
    cell_border = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side,
    )
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    headers = [
        "Order No.", "Order Type", "Status",
        "Customer Email", "Created By",
        "WL Prototype", "Fabric",
        "Style Name", "Category", "Garment Type",
        "Fit Sizes", "Total Qty", "MOQ",
        "Payment Amount", "Unit Price", "HSN/SAC Code", "GST %",
        "Customization Notes", "Message", "Fabric Type",
        "Swatch Required",
        "Admin Notes", "Enquiry No.", "Created At",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border
    ws.row_dimensions[1].height = 28

    for row_num, order in enumerate(queryset, start=2):
        ws.append([
            order.order_number,
            order.get_order_type_display(),
            order.status.replace("_", " ").title(),
            order.customer_user.email,
            order.created_by_user.email,
            order.white_label_catalogue.prototype_code if order.white_label_catalogue else "—",
            order.fabric_catalogue.fabric_name if order.fabric_catalogue else "—",
            order.style_name or "—",
            order.for_category or "—",
            order.garment_type or "—",
            order.fit_sizes or "—",
            order.total_quantity,
            order.moq or "—",
            f"₹{order.payment_amount}" if order.payment_amount else "—",
            f"₹{order.unit_price}"     if order.unit_price     else "—",
            order.hsn_code or "—",
            f"{order.gst_percentage}%" if order.gst_percentage is not None else "—",
            order.customization_notes or "—",
            order.message or "—",
            order.fabric_type or "—",
            "Yes" if order.swatch_required else "No",   # ← NEW
            order.notes or "—",
            order.enquiry.enquiry_number if order.enquiry else "—",
            order.created_at.strftime("%d %b %Y %H:%M"),
        ])
        if row_num % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = alt_fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).border = cell_border

    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 4, 16)
    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── INLINES ────────────────────────────────────────────────────────────

class OrderNoteInline(admin.TabularInline):
    model           = OrderNote
    extra           = 1
    fields          = ["note", "added_by", "created_at"]
    readonly_fields = ["added_by", "created_at"]
    ordering        = ["-created_at"]
    verbose_name    = "Note"
    verbose_name_plural = "Notes"

    def save_model(self, request, obj, form, change):
        if not obj.added_by:
            obj.added_by = request.user
        super().save_model(request, obj, form, change)


class OrderStageHistoryInline(admin.TabularInline):
    model           = OrderStageHistory
    extra           = 0
    readonly_fields = ["stage", "changed_by", "notes", "changed_at"]
    ordering        = ["changed_at"]
    can_delete      = False

    def has_add_permission(self, request, obj=None):
        return False


class OrderImageInline(admin.TabularInline):
    model           = OrderImage
    extra           = 0
    readonly_fields = ["image_preview", "file_name", "uploaded_at"]
    fields          = ["image", "image_preview", "file_name", "uploaded_at"]

    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height:80px;width:80px;object-fit:cover;border-radius:4px;" />',
                obj.image.url,
            )
        return "—"
    image_preview.short_description = "Preview"


# ── ORDER ADMIN ────────────────────────────────────────────────────────

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display  = [
        "order_number", "order_type", "customer_user",
        "assigned_to", "status", "swatch_badge",
        "total_quantity", "for_category",
        "garment_type", "payment_status_display",
        "invoice_summary", "created_at",
    ]
    list_filter   = [
        "order_type", "status",
        "swatch_required",
        "for_category", "fabric_type",
        "assigned_to",
    ]
    search_fields = ["order_number", "customer_user__email", "style_name"]
    ordering      = ["-created_at"]
    inlines       = [OrderNoteInline, OrderStageHistoryInline, OrderImageInline]
    actions       = [
        "export_all_as_excel",
        "export_wl_as_excel",
        "export_pl_as_excel",
        "export_fabrics_as_excel",
        # Swatch actions (fabrics only)
        "mark_as_swatch_sent",
        "mark_as_swatch_received",
        "mark_as_swatch_approved",
        "mark_as_swatch_rework",
        # Standard actions
        "mark_as_procurement",
        "mark_as_packing",
        "mark_as_payment_pending",
        "mark_as_payment_done",
        "mark_as_dispatch",
        "mark_as_delivered",
    ]

    readonly_fields = [
        "id", "order_number", "order_type",
        "customer_user", "created_by_user",
        "white_label_catalogue", "fabric_catalogue",
        "pl_fabric_1", "pl_fabric_2", "pl_fabric_3",
        "for_category", "garment_type", "fabric_type",
        "fit_sizes", "size_breakdown",
        "total_quantity", "moq",
        "style_name", "message",
        "swatch_required",                             # ← read-only (set by customer)
        "payment_info",
        "created_at", "updated_at",
    ]

    fieldsets = (
        ("Order Info", {
            "fields": ("id", "order_number", "order_type"),
        }),
        ("Parties", {
            "fields": ("customer_user", "created_by_user"),
            "description": "Auto-filled from the logged-in user who placed the order.",
        }),
        ("Catalogue Reference", {
            "fields": ("white_label_catalogue", "fabric_catalogue"),
        }),
        ("Order Details", {
            "fields": (
                "style_name", "for_category", "garment_type",
                "fabric_type", "fit_sizes", "size_breakdown",
                "total_quantity", "moq",
            ),
            "description": "Auto-filled at order placement.",
        }),
        ("Selected Fabrics (Private Label)", {
            "fields": ("pl_fabric_1", "pl_fabric_2", "pl_fabric_3"),
        }),
        ("Customer Notes", {
            "fields": ("customization_notes", "message"),
        }),
        # ── Swatch (fabrics orders only) ──────────────────────────── #
        ("Swatch", {
            "fields": ("swatch_required",),
            "description": (
                "Set by the customer when placing the order. "
                "If Yes → use the swatch actions below to update stage: "
                "Swatch Sent → Swatch Received → Swatch Approved → Procurement. "
                "If No → go straight to Procurement."
            ),
        }),
        # ── Admin editable ────────────────────────────────────────── #
        ("Assignment", {
            "fields": ("assigned_to",),
            "description": "Assign a staff or admin member responsible for this order.",
        }),
        ("Status", {
            "fields": ("status",),
            "description": (
                "Update order stage here. "
                "Stage history is tracked automatically on save. "
                "For fabrics with swatch: use Swatch Sent, Swatch Received, "
                "Swatch Approved stages before Procurement."
            ),
        }),
        ("Payment & Invoice", {
            "fields": (
                "payment_amount",
                "unit_price", "hsn_code", "gst_percentage",
                "payment_info",
            ),
            "description": (
                "Set the payment amount and save — a Razorpay link is auto-created when "
                "status is set to payment_pending. "
                "Unit Price, HSN/SAC Code and GST % are used to generate the tax invoice PDF. "
                "GST % is split equally as CGST + SGST (e.g. 5 = 2.5% + 2.5%)."
            ),
        }),
        ("Traceability", {
            "fields": ("enquiry",),
        }),
        ("Admin Notes", {
            "fields": ("notes",),
        }),
        ("Timestamps", {
            "fields": ("created_at", "updated_at"),
        }),
    )

    # ── Custom list display columns ────────────────────────────────── #

    def swatch_badge(self, obj):
        """Shows a badge in list view for fabric orders with swatch."""
        if obj.order_type != "fabrics":
            return mark_safe('<span style="color:#ccc;">—</span>')
        if obj.swatch_required:
            return mark_safe(
                '<span style="background:#f0e6ff;color:#5b21b6;padding:2px 8px;'
                'border-radius:10px;font-size:11px;font-weight:600;">Swatch</span>'
            )
        return mark_safe(
            '<span style="background:#f0f0f0;color:#666;padding:2px 8px;'
            'border-radius:10px;font-size:11px;">Direct</span>'
        )
    swatch_badge.short_description = "Swatch"

    def invoice_summary(self, obj):
        """Shows unit price + GST in the list view when set."""
        if not obj.unit_price:
            return mark_safe('<span style="color:#ccc;">—</span>')
        from decimal import Decimal, ROUND_HALF_UP
        qty      = Decimal(str(obj.total_quantity))
        rate     = Decimal(str(obj.unit_price))
        gst_pct  = Decimal(str(obj.gst_percentage or 5))
        subtotal = (rate * qty).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        half     = gst_pct / 2
        cgst     = (subtotal * half / 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        sgst     = cgst
        total    = (subtotal + cgst + sgst).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        return format_html(
            '<span style="font-size:11px;color:#333;">'
            '₹{}/u &nbsp;|&nbsp; GST {}% &nbsp;|&nbsp; <b>₹{}</b>'
            '</span>',
            obj.unit_price, gst_pct, total,
        )
    invoice_summary.short_description = "Invoice"

    def payment_status_display(self, obj):
        from payments.models import PaymentTransaction, PaymentStatus
        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(obj)
        tx = PaymentTransaction.objects.filter(
            content_type=ct, object_id=obj.id
        ).order_by("-created_at").first()

        if not tx:
            return mark_safe('<span style="color:#999;">—</span>')

        colors = {
            "pending":  "#f0a500",
            "paid":     "#27ae60",
            "failed":   "#e74c3c",
            "refunded": "#8e44ad",
        }
        color = colors.get(tx.status, "#999")
        return format_html(
            '<span style="color:{};font-weight:600;">{}</span>',
            color, tx.get_status_display()
        )
    payment_status_display.short_description = "Payment"

    def payment_info(self, obj):
        from payments.models import PaymentTransaction
        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(obj)
        tx = PaymentTransaction.objects.filter(
            content_type=ct, object_id=obj.id
        ).order_by("-created_at").first()

        if not tx:
            return mark_safe('<span style="color:#999;">No payment created yet.</span>')

        color_map = {
            "pending":  "#f0a500",
            "paid":     "#27ae60",
            "failed":   "#e74c3c",
            "refunded": "#8e44ad",
        }
        color = color_map.get(tx.status, "#999")
        html = f"""
        <table style="border-collapse:collapse;font-size:13px;">
          <tr><td style="padding:3px 10px 3px 0;color:#666;">Status</td>
              <td style="color:{color};font-weight:600;">{tx.get_status_display()}</td></tr>
          <tr><td style="padding:3px 10px 3px 0;color:#666;">Amount</td>
              <td>₹{tx.amount} {tx.currency}</td></tr>
          <tr><td style="padding:3px 10px 3px 0;color:#666;">Razorpay Order ID</td>
              <td>{tx.razorpay_order_id or '—'}</td></tr>
          <tr><td style="padding:3px 10px 3px 0;color:#666;">Payment Reference</td>
              <td>{tx.payment_reference or '—'}</td></tr>
          <tr><td style="padding:3px 10px 3px 0;color:#666;">Paid At</td>
              <td>{tx.paid_at.strftime('%d %b %Y %H:%M') if tx.paid_at else '—'}</td></tr>
        </table>
        """
        return mark_safe(html)
    payment_info.short_description = "Payment Details"

    # ── Save logic ─────────────────────────────────────────────────── #

    def save_model(self, request, obj, form, change):
        if change:
            old = Order.objects.get(pk=obj.pk)

            # Track status change in stage history
            if old.status != obj.status:
                OrderStageHistory.objects.create(
                    order      = obj,
                    stage      = obj.status,
                    changed_by = request.user,
                    notes      = "Updated via admin panel.",
                )
                # Send push notification to customer
                try:
                    from notifications.service import send_order_stage_notification
                    send_order_stage_notification(obj, obj.status)
                except Exception as e:
                    import logging
                    logging.getLogger(__name__).error(
                        f"Notification failed for {obj.order_number}: {e}"
                    )

            # Notify staff when assigned_to changes
            old_assignee = old.assigned_to_id
            new_assignee = obj.assigned_to
            if new_assignee and old_assignee != new_assignee.pk:
                try:
                    from notifications.service import send_order_assigned_notification
                    send_order_assigned_notification(
                        order            = obj,
                        assigned_to_user = new_assignee,
                        assigned_by_user = request.user,
                    )
                except Exception as e:
                    import logging
                    logging.getLogger(__name__).error(
                        f"Assignment notification failed for {obj.order_number}: {e}"
                    )

            # Auto-create Razorpay payment when status → payment_pending
            status_just_changed = (old.status != "payment_pending" and obj.status == "payment_pending")
            amount_just_set     = (obj.status == "payment_pending" and obj.payment_amount and not old.payment_amount)

            if (status_just_changed or amount_just_set) and obj.payment_amount:
                self._create_razorpay_payment(request, obj)

        super().save_model(request, obj, form, change)

    def _create_razorpay_payment(self, request, order):
        import traceback
        try:
            from payments import gateway
            from payments.models import PaymentTransaction, PaymentStatus
            from django.contrib.contenttypes.models import ContentType

            messages.info(
                request,
                f"Creating payment for {order.order_number} | Amount: ₹{order.payment_amount}"
            )

            ct = ContentType.objects.get_for_model(order)
            existing = PaymentTransaction.objects.filter(
                content_type=ct,
                object_id=order.id,
                status=PaymentStatus.PENDING,
            )
            if existing.exists():
                messages.warning(
                    request,
                    f"Pending payment already exists: {existing.first().razorpay_order_id}"
                )
                return

            result = gateway.create_payment(
                content_object = order,
                amount         = order.payment_amount,
                payment_type   = "order",
                paid_by        = order.customer_user,
                notes          = f"Payment for {order.order_number}",
            )
            messages.success(
                request,
                f"Razorpay payment created. "
                f"Order ID: {result['razorpay_order_id']} | Amount: ₹{order.payment_amount}"
            )
        except Exception as e:
            messages.error(request, f"Payment error: {str(e)}")
            messages.error(request, f"Traceback: {traceback.format_exc()[:600]}")

    # ── Bulk actions ───────────────────────────────────────────────── #

    @admin.action(description="Export selected orders to Excel")
    def export_all_as_excel(self, request, queryset):
        qs = queryset.select_related(
            "customer_user", "created_by_user",
            "white_label_catalogue", "fabric_catalogue", "enquiry",
        )
        return export_orders_to_excel(
            qs, f"orders_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    @admin.action(description="Export White Label orders to Excel")
    def export_wl_as_excel(self, request, queryset):
        qs = queryset.filter(order_type="white_label").select_related(
            "customer_user", "created_by_user", "white_label_catalogue", "enquiry",
        )
        return export_orders_to_excel(qs, f"orders_wl_{timezone.now().strftime('%Y%m%d')}.xlsx")

    @admin.action(description="Export Private Label orders to Excel")
    def export_pl_as_excel(self, request, queryset):
        qs = queryset.filter(order_type="private_label").select_related(
            "customer_user", "created_by_user", "enquiry",
        )
        return export_orders_to_excel(qs, f"orders_pl_{timezone.now().strftime('%Y%m%d')}.xlsx")

    @admin.action(description="Export Fabrics orders to Excel")
    def export_fabrics_as_excel(self, request, queryset):
        qs = queryset.filter(order_type="fabrics").select_related(
            "customer_user", "created_by_user", "fabric_catalogue", "enquiry",
        )
        return export_orders_to_excel(
            qs, f"orders_fabrics_{timezone.now().strftime('%Y%m%d')}.xlsx")

    # ── Swatch stage actions (fabrics orders) ──────────────────────── #

    @admin.action(description="[Fabrics] Mark as Swatch Sent")
    def mark_as_swatch_sent(self, request, queryset):
        self._bulk_update_status(request, queryset, "swatch_sent",
                                  require_swatch=True)

    @admin.action(description="[Fabrics] Mark as Swatch Received")
    def mark_as_swatch_received(self, request, queryset):
        self._bulk_update_status(request, queryset, "swatch_received",
                                  require_swatch=True)

    @admin.action(description="[Fabrics] Mark as Swatch Approved")
    def mark_as_swatch_approved(self, request, queryset):
        self._bulk_update_status(request, queryset, "swatch_approved",
                                  require_swatch=True)

    @admin.action(description="[Fabrics] Mark as Swatch Rework")
    def mark_as_swatch_rework(self, request, queryset):
        self._bulk_update_status(request, queryset, "swatch_rework",
                                  require_swatch=True)

    # ── Standard stage actions ─────────────────────────────────────── #

    @admin.action(description="Mark selected as Procurement")
    def mark_as_procurement(self, request, queryset):
        self._bulk_update_status(request, queryset, "procurement")

    @admin.action(description="Mark selected as Packing")
    def mark_as_packing(self, request, queryset):
        self._bulk_update_status(request, queryset, "packing")

    @admin.action(description="Mark selected as Payment Pending")
    def mark_as_payment_pending(self, request, queryset):
        self._bulk_update_status(request, queryset, "payment_pending")

    @admin.action(description="Mark selected as Payment Done")
    def mark_as_payment_done(self, request, queryset):
        self._bulk_update_status(request, queryset, "payment_done")

    @admin.action(description="Mark selected as Dispatched")
    def mark_as_dispatch(self, request, queryset):
        self._bulk_update_status(request, queryset, "dispatch")

    @admin.action(description="Mark selected as Delivered")
    def mark_as_delivered(self, request, queryset):
        self._bulk_update_status(request, queryset, "delivered")

    def _bulk_update_status(self, request, queryset, new_status,
                             require_swatch=False):
        updated  = 0
        skipped  = 0
        for order in queryset:
            # Skip if swatch action applied to non-swatch order
            if require_swatch and not order.swatch_required:
                skipped += 1
                continue
            # Skip if status already set
            if order.status == new_status:
                continue
            order.status = new_status
            order.save(update_fields=["status", "updated_at"])
            OrderStageHistory.objects.create(
                order      = order,
                stage      = new_status,
                changed_by = request.user,
                notes      = "Bulk updated via admin panel.",
            )
            # Send push notification to customer
            try:
                from notifications.service import send_order_stage_notification
                send_order_stage_notification(order, new_status)
            except Exception as e:
                import logging
                logging.getLogger(__name__).error(
                    f"Notification failed for {order.order_number}: {e}"
                )
            updated += 1

        label = new_status.replace("_", " ").title()
        self.message_user(
            request,
            f"{updated} order(s) updated to '{label}'."
            + (f" {skipped} skipped (swatch not requested)." if skipped else ""),
        )

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        User = get_user_model()
        if "assigned_to" in form.base_fields:
            form.base_fields["assigned_to"].queryset = User.objects.filter(
                role__in=["admin", "staff"], is_active=True
            ).order_by("email")
        return form

    def has_add_permission(self, request):
        return False


# ── STAGE HISTORY ADMIN ────────────────────────────────────────────────

@admin.register(OrderStageHistory)
class OrderStageHistoryAdmin(admin.ModelAdmin):
    list_display    = ["order", "stage", "changed_by", "notes", "changed_at"]
    list_filter     = ["stage"]
    search_fields   = ["order__order_number"]
    readonly_fields = ["id", "order", "stage", "changed_by", "changed_at"]
    ordering        = ["-changed_at"]

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False


# ── ORDER IMAGE ADMIN ──────────────────────────────────────────────────

@admin.register(OrderImage)
class OrderImageAdmin(admin.ModelAdmin):
    list_display    = ["file_name", "order", "image_preview", "uploaded_at"]
    search_fields   = ["file_name", "order__order_number"]
    readonly_fields = ["id", "image_preview", "order", "file_name", "uploaded_at"]
    ordering        = ["-uploaded_at"]

    def image_preview(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" style="height:80px;width:80px;object-fit:cover;border-radius:4px;" />',
                obj.image.url,
            )
        return "—"
    image_preview.short_description = "Preview"

    def has_add_permission(self, request):
        return False