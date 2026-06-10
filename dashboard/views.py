import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from accounts.permissions import IsAdminOrStaff


# ── HELPERS ────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")
BORDER_SIDE   = Side(style="thin", color="CCCCCC")
CELL_BORDER   = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)


def style_header_row(ws, headers):
    """Apply header styling to row 1."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = CELL_BORDER


def style_data_rows(ws, start_row, end_row, num_cols):
    """Apply alternating row colours and borders."""
    for row in range(start_row, end_row + 1):
        fill = ALT_FILL if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill


def auto_fit_columns(ws, headers):
    """Set column widths based on header length."""
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(header) + 4, 16)
    ws.row_dimensions[1].height = 30


def make_response(wb, filename):
    """Return Excel file as HTTP response."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def fmt_dt(dt):
    """Format datetime nicely."""
    if not dt:
        return "—"
    return dt.strftime("%d %b %Y %H:%M")


def fmt_date(d):
    """Format date nicely."""
    if not d:
        return "—"
    return d.strftime("%d %b %Y")


# ── EXPORT ENQUIRIES ───────────────────────────────────────────────────

class ExportEnquiriesView(APIView):
    """
    GET /api/dashboard/export/enquiries/
    Admin/Staff — exports all enquiries to Excel.

    Optional filters:
      ?order_type=private_label | white_label | fabrics | others
      ?status=new | contacted | ...
      ?date_from=2026-01-01
      ?date_to=2026-03-31
    """
    permission_classes = [IsAuthenticated, IsAdminOrStaff]

    def get(self, request):
        from enquiries.models import Enquiry

        qs = Enquiry.objects.select_related(
            "assigned_to_user", "wl_prototype", "fabric", "customer"
        ).order_by("-created_at")

        # Apply filters
        order_type = request.query_params.get("order_type")
        status     = request.query_params.get("status")
        date_from  = request.query_params.get("date_from")
        date_to    = request.query_params.get("date_to")

        if order_type: qs = qs.filter(order_type=order_type)
        if status:     qs = qs.filter(status=status)
        if date_from:  qs = qs.filter(created_at__date__gte=date_from)
        if date_to:    qs = qs.filter(created_at__date__lte=date_to)

        wb = openpyxl.Workbook()

        # ── Sheet 1: All Enquiries ─────────────────────────────────── #
        ws = wb.active
        ws.title = "All Enquiries"

        headers = [
            "Enquiry No.", "Order Type", "Status", "Source Page",
            "Full Name", "Phone", "Email", "Brand Name",
            "Company Age (Yrs)", "Total Pieces Req.", "Annual Revenue",
            "Message",
            "WL Prototype Code", "Fabric Name",
            "Assigned To", "Is Viewed", "Viewed At",
            "Admin Notes", "Created At", "Updated At",
        ]

        style_header_row(ws, headers)

        for row_num, enq in enumerate(qs, start=2):
            ws.append([
                enq.enquiry_number,
                enq.get_order_type_display(),
                enq.get_status_display(),
                enq.get_source_page_display() if enq.source_page else "—",
                enq.full_name,
                enq.phone,
                enq.email,
                enq.brand_name,
                enq.company_age_years or "—",
                enq.total_pieces_required or "—",
                enq.annual_revenue or "—",
                enq.message,
                enq.wl_prototype.prototype_code if enq.wl_prototype else "—",
                enq.fabric.fabric_name if enq.fabric else "—",
                enq.assigned_to_user.email if enq.assigned_to_user else "—",
                "Yes" if enq.is_viewed else "No",
                fmt_dt(enq.viewed_at),
                enq.admin_notes or "—",
                fmt_dt(enq.created_at),
                fmt_dt(enq.updated_at),
            ])

        style_data_rows(ws, 2, qs.count() + 1, len(headers))
        auto_fit_columns(ws, headers)
        ws.freeze_panes = "A2"

        # ── Sheet 2: Summary by Order Type ────────────────────────── #
        ws2 = wb.create_sheet("Summary by Type")
        summary_headers = ["Order Type", "Total", "New", "Contacted",
                            "Prospect", "Accepted", "Rejected", "Closed"]
        style_header_row(ws2, summary_headers)

        order_types = ["private_label", "white_label", "fabrics", "others"]
        for i, ot in enumerate(order_types, start=2):
            subset = qs.filter(order_type=ot)
            ws2.append([
                ot.replace("_", " ").title(),
                subset.count(),
                subset.filter(status="new").count(),
                subset.filter(status="contacted").count(),
                subset.filter(status="prospect").count(),
                subset.filter(status="accepted").count(),
                subset.filter(status="rejected").count(),
                subset.filter(status="closed").count(),
            ])

        style_data_rows(ws2, 2, len(order_types) + 1, len(summary_headers))
        auto_fit_columns(ws2, summary_headers)

        filename = f"enquiries_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return make_response(wb, filename)


# ── EXPORT ORDERS ──────────────────────────────────────────────────────

class ExportOrdersView(APIView):
    """
    GET /api/dashboard/export/orders/
    Admin/Staff — exports all orders to Excel.

    Optional filters:
      ?order_type=private_label | white_label | fabrics
      ?status=order_placed | cutting | ...
      ?date_from=2026-01-01
      ?date_to=2026-03-31
    """
    permission_classes = [IsAuthenticated, IsAdminOrStaff]

    def get(self, request):
        from orders.models import Order

        qs = Order.objects.select_related(
            "customer_user", "created_by_user",
            "white_label_catalogue", "fabric_catalogue",
            "enquiry",
        ).order_by("-created_at")

        # Apply filters
        order_type = request.query_params.get("order_type")
        status     = request.query_params.get("status")
        date_from  = request.query_params.get("date_from")
        date_to    = request.query_params.get("date_to")

        if order_type: qs = qs.filter(order_type=order_type)
        if status:     qs = qs.filter(status=status)
        if date_from:  qs = qs.filter(created_at__date__gte=date_from)
        if date_to:    qs = qs.filter(created_at__date__lte=date_to)

        wb = openpyxl.Workbook()

        # ── Sheet 1: All Orders ────────────────────────────────────── #
        ws = wb.active
        ws.title = "All Orders"

        headers = [
            "Order No.", "Order Type", "Status",
            "Customer Email", "Created By",
            "WL Prototype", "Fabric",
            "Style Name", "Category", "Garment Type",
            "Fit Sizes", "Total Qty", "MOQ",
            "Customization Notes", "Message", "Fabric Type",
            "Payment Amount", "Admin Notes",
            "Enquiry No.", "Created At", "Updated At",
        ]

        style_header_row(ws, headers)

        for row_num, order in enumerate(qs, start=2):
            ws.append([
                order.order_number,
                order.get_order_type_display(),
                order.status.replace("_", " ").title(),
                order.customer_user.email,
                order.created_by_user.email,
                order.white_label_catalogue.prototype_code if order.white_label_catalogue else "—",
                order.fabric_catalogue.fabric_name if order.fabric_catalogue else "—",
                order.style_name or "—",
                order.for_category or "—",
                order.garment_type or "—",
                order.fit_sizes or "—",
                order.total_quantity,
                order.moq or "—",
                order.customization_notes or "—",
                order.message or "—",
                order.fabric_type or "—",
                f"₹{order.payment_amount}" if order.payment_amount else "—",
                order.notes or "—",
                order.enquiry.enquiry_number if order.enquiry else "—",
                fmt_dt(order.created_at),
                fmt_dt(order.updated_at),
            ])

        style_data_rows(ws, 2, qs.count() + 1, len(headers))
        auto_fit_columns(ws, headers)
        ws.freeze_panes = "A2"

        # ── Sheet 2: Summary by Order Type ────────────────────────── #
        ws2 = wb.create_sheet("Summary by Type")
        summary_headers = [
            "Order Type", "Total",
            "Order Placed", "In Progress", "Payment Pending",
            "Payment Done", "Dispatched", "Delivered",
        ]
        style_header_row(ws2, summary_headers)

        order_types = ["private_label", "white_label", "fabrics"]
        for ot in order_types:
            subset = qs.filter(order_type=ot)
            in_progress = subset.exclude(
                status__in=["order_placed", "payment_pending",
                            "payment_done", "dispatch", "delivered"]
            ).count()
            ws2.append([
                ot.replace("_", " ").title(),
                subset.count(),
                subset.filter(status="order_placed").count(),
                in_progress,
                subset.filter(status="payment_pending").count(),
                subset.filter(status="payment_done").count(),
                subset.filter(status="dispatch").count(),
                subset.filter(status="delivered").count(),
            ])

        style_data_rows(ws2, 2, len(order_types) + 1, len(summary_headers))
        auto_fit_columns(ws2, summary_headers)

        # ── Sheet 3: Stage History ─────────────────────────────────── #
        ws3 = wb.create_sheet("Stage History")
        history_headers = [
            "Order No.", "Order Type", "Stage",
            "Changed By", "Notes", "Changed At",
        ]
        style_header_row(ws3, history_headers)

        from orders.models import OrderStageHistory
        history_qs = OrderStageHistory.objects.select_related(
            "order", "changed_by"
        ).filter(order__in=qs).order_by("order__order_number", "changed_at")

        for row_num, h in enumerate(history_qs, start=2):
            ws3.append([
                h.order.order_number,
                h.order.get_order_type_display(),
                h.stage.replace("_", " ").title(),
                h.changed_by.email if h.changed_by else "System",
                h.notes or "—",
                fmt_dt(h.changed_at),
            ])

        style_data_rows(ws3, 2, history_qs.count() + 1, len(history_headers))
        auto_fit_columns(ws3, history_headers)

        filename = f"orders_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return make_response(wb, filename)


# ── DASHBOARD SUMMARY ──────────────────────────────────────────────────

class DashboardSummaryView(APIView):
    """
    GET /api/dashboard/summary/
    Admin/Staff — returns counts for dashboard widgets.
    """
    permission_classes = [IsAuthenticated, IsAdminOrStaff]

    def get(self, request):
        from enquiries.models import Enquiry
        from orders.models import Order
        from django.db.models import Sum, Value
        from django.db.models.functions import Coalesce
        from collections import defaultdict

        # Single query per model using conditional aggregation
        enq = Enquiry.objects.aggregate(
            total         = Count("id"),
            unread        = Count("id", filter=Q(is_viewed=False)),
            new           = Count("id", filter=Q(status="new")),
            private_label = Count("id", filter=Q(order_type="private_label")),
            white_label   = Count("id", filter=Q(order_type="white_label")),
            fabrics       = Count("id", filter=Q(order_type="fabrics")),
            others        = Count("id", filter=Q(order_type="others")),
        )

        ord = Order.objects.aggregate(
            total           = Count("id"),
            order_placed    = Count("id", filter=Q(status="order_placed")),
            payment_pending = Count("id", filter=Q(status="payment_pending")),
            payment_done    = Count("id", filter=Q(status="payment_done")),
            dispatched      = Count("id", filter=Q(status="dispatch")),
            delivered       = Count("id", filter=Q(status="delivered")),
            private_label   = Count("id", filter=Q(order_type="private_label")),
            white_label     = Count("id", filter=Q(order_type="white_label")),
            fabrics         = Count("id", filter=Q(order_type="fabrics")),
        )
        in_progress = (
            ord["total"]
            - ord["order_placed"]
            - ord["payment_pending"]
            - ord["payment_done"]
            - ord["dispatched"]
            - ord["delivered"]
        )

        # Stage-wise stats aggregation
        pipeline_orders = Order.objects.filter(assigned_to=request.user)

        state_raw = pipeline_orders.values(
            'status',
            'order_type'
        ).annotate(
            count=Count('id'),
            total_qty=Coalesce(Sum('total_quantity'), Value(0))
        )

        STATUS_PIPELINE = [
            "order_placed",
            "sample_request",
            "sample_approval",
            "sample_rework",
            "swatch_sent",
            "swatch_received",
            "swatch_approved",
            "swatch_rework",
            "order_confirmed",
            "advance_pending",
            "advance_paid",
            "bulk_production",
            "quality_inspection",
            "packing",
            "payment_pending",
            "payment_done",
            "dispatch",
            "shipment_tracking",
            "delivered",
            "rework_replacement_pending",
            "order_completed",
            "cancelled",
        ]
        STATUS_ORDER_MAP = {status: index for index, status in enumerate(STATUS_PIPELINE)}

        STATUS_DISPLAY_MAP = {
            "order_placed": "Order Placed",
            "sample_request": "Sample Request",
            "sample_approval": "Sample Approved",
            "sample_rework": "Sample Rework",
            "swatch_sent": "Swatch Sent",
            "swatch_received": "Swatch Received",
            "swatch_approved": "Swatch Approved",
            "swatch_rework": "Swatch Rework",
            "order_confirmed": "Order Confirmed",
            "advance_pending": "Advance Pending",
            "advance_paid": "Advance Paid",
            "bulk_production": "Bulk Production",
            "quality_inspection": "Quality Inspection",
            "packing": "Packing",
            "payment_pending": "Payment Pending",
            "payment_done": "Payment Done",
            "dispatch": "Dispatch",
            "shipment_tracking": "Shipment",
            "delivered": "Delivered",
            "rework_replacement_pending": "Rework / Replacement Pending",
            "order_completed": "Order Completed",
            "cancelled": "Cancelled",
        }

        state_totals = defaultdict(lambda: {
            "wl_orders": 0,
            "wl_pieces": 0,
            "pl_orders": 0,
            "pl_pieces": 0,
            "combined_orders": 0,
            "combined_pieces": 0,
            "fabrics_orders": 0,
            "fabrics_meters": 0,
        })

        for entry in state_raw:
            status_key = entry['status']
            if not status_key or not status_key.strip():
                status_key = 'order_placed'
                
            stats = state_totals[status_key]
            ot = entry['order_type']
            count = entry['count']
            qty = entry['total_qty']
            
            if ot == 'white_label':
                stats["wl_orders"] += count
                stats["wl_pieces"] += qty
                stats["combined_orders"] += count
                stats["combined_pieces"] += qty
            elif ot == 'private_label':
                stats["pl_orders"] += count
                stats["pl_pieces"] += qty
                stats["combined_orders"] += count
                stats["combined_pieces"] += qty
            elif ot == 'fabrics':
                stats["fabrics_orders"] += count
                stats["fabrics_meters"] += qty

        state_wise_list = []
        sorted_status_keys = sorted(state_totals.keys(), key=lambda k: STATUS_ORDER_MAP.get(k, 999))
        for status_key in sorted_status_keys:
            stats = state_totals[status_key]
            display_name = STATUS_DISPLAY_MAP.get(status_key, status_key.replace('_', ' ').title())
            state_wise_list.append({
                "state": status_key,
                "display_name": display_name,
                **stats
            })

        return Response({
            "enquiries": {
                "total": enq["total"],
                "unread": enq["unread"],
                "new": enq["new"],
                "by_type": {
                    "private_label": enq["private_label"],
                    "white_label":   enq["white_label"],
                    "fabrics":       enq["fabrics"],
                    "others":        enq["others"],
                },
            },
            "orders": {
                "total":           ord["total"],
                "order_placed":    ord["order_placed"],
                "in_progress":     in_progress,
                "payment_pending": ord["payment_pending"],
                "payment_done":    ord["payment_done"],
                "dispatched":      ord["dispatched"],
                "delivered":       ord["delivered"],
                "by_type": {
                    "private_label": ord["private_label"],
                    "white_label":   ord["white_label"],
                    "fabrics":       ord["fabrics"],
                },
            },
            "state_wise_stats": state_wise_list,
        })


def dashboard_callback(request, context):
    """
    Context callback for the Django Unfold admin dashboard.
    Populates context with statistics and recent activity lists.
    """
    from enquiries.models import Enquiry
    from orders.models import Order
    from accounts.models import User
    from payments.models import PaymentTransaction
    from django.db.models import Sum, Count, Q
    from datetime import datetime
    import calendar
    from django.utils import timezone

    if not request.user or not request.user.is_authenticated:
        return context

    if request.user.role == "vendor":
        from catalogue.models import WLPrototype, FabricsCatalogue

        wl_stats = WLPrototype.objects.filter(created_by_admin=request.user).aggregate(
            total=Count("id"),
            active=Count("id", filter=Q(is_active=True)),
        )

        fabric_stats = FabricsCatalogue.objects.filter(created_by=request.user).aggregate(
            total=Count("id"),
            active=Count("id", filter=Q(is_active=True)),
        )

        wl_total = wl_stats["total"] or 0
        wl_active = wl_stats["active"] or 0
        wl_inactive = wl_total - wl_active

        fabric_total = fabric_stats["total"] or 0
        fabric_active = fabric_stats["active"] or 0
        fabric_inactive = fabric_total - fabric_active

        recent_wl = WLPrototype.objects.filter(created_by_admin=request.user).order_by("-created_at")[:5]
        recent_fabrics = FabricsCatalogue.objects.filter(created_by=request.user).order_by("-created_at")[:5]

        context.update({
            "vendor_stats": {
                "wl_total": wl_total,
                "wl_active": wl_active,
                "wl_inactive": wl_inactive,
                "fabric_total": fabric_total,
                "fabric_active": fabric_active,
                "fabric_inactive": fabric_inactive,
                "total_products": wl_total + fabric_total,
                "total_active": wl_active + fabric_active,
                "total_inactive": wl_inactive + fabric_inactive,
            },
            "recent_wl_products": recent_wl,
            "recent_fabric_products": recent_fabrics,
        })
        return context

    # Only superusers get custom dashboard stats/charts context
    if not request.user.is_superuser:
        return context

    # Query statistics
    enq_stats = Enquiry.objects.aggregate(
        total=Count("id"),
        unread=Count("id", filter=Q(is_viewed=False)),
        new=Count("id", filter=Q(status="new")),
    )
    
    ord_stats = Order.objects.aggregate(
        total=Count("id"),
        order_placed=Count("id", filter=Q(status="order_placed")),
        payment_pending=Count("id", filter=Q(status="payment_pending")),
        payment_done=Count("id", filter=Q(status="payment_done")),
        delivered=Count("id", filter=Q(status="delivered")),
        private_label=Count("id", filter=Q(order_type="private_label")),
        white_label=Count("id", filter=Q(order_type="white_label")),
        fabrics=Count("id", filter=Q(order_type="fabrics")),
        total_revenue=Sum("total_amount", filter=Q(status__in=[
            "advance_paid", "bulk_production", "quality_inspection", "packing",
            "payment_done", "dispatch", "shipment_tracking", "delivered", "order_completed"
        ]))
    )
    
    user_stats = User.objects.aggregate(
        total_customers=Count("id", filter=Q(role="customer")),
        total_vendors=Count("id", filter=Q(role="vendor")),
    )

    # In-progress calculation (all statuses that aren't placed, payment pending/done, delivered, or cancelled)
    completed_or_pending_stages = ["order_placed", "payment_pending", "payment_done", "delivered", "cancelled", "order_completed"]
    in_progress_count = Order.objects.exclude(status__in=completed_or_pending_stages).count()

    # 1. Calculate Last 6 Months Revenue Trend (Database-Agnostic)
    revenue_trend = []
    today = timezone.now().date()
    for i in range(5, -1, -1):
        year = today.year
        month = today.month - i
        if month <= 0:
            month += 12
            year -= 1
        
        start_date = datetime(year, month, 1, 0, 0, 0)
        _, last_day = calendar.monthrange(year, month)
        end_date = datetime(year, month, last_day, 23, 59, 59)
        
        # Filter paid/delivered orders for revenue calculation
        monthly_sum = Order.objects.filter(
            created_at__range=(start_date, end_date),
            status__in=[
                "advance_paid", "bulk_production", "quality_inspection", "packing",
                "payment_done", "dispatch", "shipment_tracking", "delivered", "order_completed"
            ]
        ).aggregate(total=Sum("total_amount"))["total"] or 0
        
        month_name = calendar.month_abbr[month] + f" '{str(year)[2:]}"
        revenue_trend.append({
            "month": month_name,
            "revenue": float(monthly_sum)
        })

    # 2. Calculate Enquiry Status Distribution
    enq_status_counts = Enquiry.objects.values('status').annotate(count=Count('id'))
    status_map = {item['status']: item['count'] for item in enq_status_counts}
    enquiry_statuses = {
        "New": status_map.get("new", 0),
        "Contacted": status_map.get("contacted", 0),
        "Prospect": status_map.get("prospect", 0),
        "Accepted": status_map.get("accepted", 0),
        "Closed": status_map.get("closed", 0),
    }

    # Recent lists
    recent_enquiries = Enquiry.objects.select_related("assigned_to_user").order_by("-created_at")[:5]
    recent_orders = Order.objects.select_related("customer_user").order_by("-created_at")[:5]
    
    # Generic relation helper to fetch content types for transactions
    from django.contrib.contenttypes.models import ContentType
    order_ct = ContentType.objects.get_for_model(Order)
    recent_payments = PaymentTransaction.objects.select_related("paid_by").order_by("-created_at")[:5]

    context.update({
        "stats": {
            "enquiries_total": enq_stats["total"],
            "enquiries_unread": enq_stats["unread"],
            "enquiries_new": enq_stats["new"],
            "orders_total": ord_stats["total"],
            "orders_placed": ord_stats["order_placed"],
            "orders_in_progress": in_progress_count,
            "orders_payment_pending": ord_stats["payment_pending"],
            "orders_payment_done": ord_stats["payment_done"],
            "orders_delivered": ord_stats["delivered"],
            "total_revenue": ord_stats["total_revenue"] or 0,
            "this_month_revenue": revenue_trend[-1]["revenue"] if revenue_trend else 0,
            "total_customers": user_stats["total_customers"],
            "total_vendors": user_stats["total_vendors"],
        },
        "recent_enquiries": recent_enquiries,
        "recent_orders": recent_orders,
        "recent_payments": recent_payments,
        "charts": {
            "revenue_months": [r["month"] for r in revenue_trend],
            "revenue_values": [r["revenue"] for r in revenue_trend],
            "order_types": ["Private Label", "White Label", "Fabrics"],
            "order_values": [
                ord_stats["private_label"] or 0,
                ord_stats["white_label"] or 0,
                ord_stats["fabrics"] or 0
            ],
            "enquiry_status_labels": list(enquiry_statuses.keys()),
            "enquiry_status_values": list(enquiry_statuses.values()),
        }
    })
    return context